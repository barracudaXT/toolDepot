import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import logging
import openpyxl
from openpyxl import Workbook
import dns.resolver
import os
import threading
import progressbar

# Setting up logging to monitor performance and errors
logging.basicConfig(level=logging.INFO)

class CnameRetrievalApp:
    def __init__(self, root):
        self.root = root
        self.root.title("CNAME RETRIEVER")
        self.root.geometry("280x250")

        self.source_label = tk.Label(root, text="Choose the excel file:",font = ("Arial", 12))
        self.source_label.pack(pady=5)

        self.source_path = tk.StringVar()
        self.source_entry = tk.Entry(root, width = 40, textvariable=self.source_path)
        self.source_entry.pack(pady=5)

        self.browse_button = tk.Button(root, width = 10, text="Browse", font = ("Arial", 10), command=self.browse_source)
        self.browse_button.pack(pady=5)

        self.progress_bar = tk.ttk.Progressbar(root, orient="horizontal", length=200, mode="determinate")
        self.progress_bar.pack(pady=10)

        self.retrieve_button = tk.Button(root, width = 20, text="Retrieve CNAME", font = ("Arial", 10), command=self.retrieve_cname)
        self.retrieve_button.pack(pady=5)

        self.reminder_label = tk.Label(root, text="== Output file is saved at == \n == the same location as the source file. ==")
        self.reminder_label.pack(pady=10)

    def browse_source(self):
        file_path = filedialog.askopenfilename(filetypes=[("XLSX files", "*.xlsx")])
        self.source_path.set(file_path)

    def retrieve_cname(self):
        source_file = self.source_path.get()

        if not source_file:
            messagebox.showerror("Error", "Please provide the source file path.")
            return

        output_file = os.path.join(os.path.dirname(source_file), "Retrieved_CNAME.xlsx")

        records = self.retrieve_cname_records(source_file)
        if records:
            self.store_results_to_xlsx(records, output_file)
            messagebox.showinfo("Success", "CNAME records retrieved and stored successfully.")
        else:
            messagebox.showwarning("Warning", "Failed to retrieve CNAME records from the XLSX file.")

    def update_progress_bar(self, value):
        self.progress_bar["value"] = value
        self.root.update_idletasks()

    def retrieve_cname_records(self, file_path):
        try:
            logging.info("Reading XLSX file...")
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            logging.info("Retrieving CNAME records...")
            records = {}

            total_rows = sheet.max_row - 1
            current_row = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                website = row[0]
                try:
                    cname = dns.resolver.resolve(website, 'CNAME')[0].target.to_text()
                    records[website] = cname

                except dns.resolver.NoAnswer:
                    logging.warning(f"No CNAME record found for {website}")
                except dns.resolver.NXDOMAIN:
                    logging.warning(f"{website} does not exist")

                current_row += 1
                progress_percentage = (current_row / total_rows) * 100
                self.update_progress_bar(progress_percentage)

            return records

        except Exception as e:
            logging.error(f"An error occurred: {e}")
            return {}

    def store_results_to_xlsx(self, records, output_path):
        try:
            logging.info("Storing results in xlsx file...")
            workbook = Workbook()
            sheet = workbook.active

            for i, (website, cname) in enumerate(records.items(), start=1):
                sheet.cell(row=i, column=1, value=website)
                sheet.cell(row=i, column=2, value=cname)

            workbook.save(output_path)
            logging.info("Results stored successfully in %s", output_path)
        except Exception as e:
            logging.error(f"An error occurred: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = CnameRetrievalApp(root)
    root.mainloop()
